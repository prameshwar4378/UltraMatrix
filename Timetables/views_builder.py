import json
import re
from collections import defaultdict
from io import BytesIO
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.utils import timezone

from Academic.models import AcademicYear, BellSchedule, Day, Period
from Timetables.models import (
    LectureAdjustment,
    TeacherDailyStatus,
    Timetable,
    TimetableConfiguration,
    LessonAllocation,
    ClassSection,
)
from Teachers.models import Teacher, TeacherAvailability
from Subjects.models import Subject, TeacherSubjectCapability
from Rooms.models import Room
from Accounts.utils import get_current_school, get_school_object_or_404
import json
from django.views.decorators.csrf import csrf_exempt
from AI_TIMETABLE_SAAS.logging_utils import log_exceptions
from .readiness import timetable_readiness

@login_required
@log_exceptions
def timetable_builder(request, template_name="timetable_builder.html"):
    current_school = get_current_school(request)
    if not current_school:
        return render(request, template_name, {
            "academic_years": AcademicYear.objects.none(),
            "timetables": Timetable.objects.none(),
            "selected_academic_year_id": "",
            "selected_timetable_id": "",
            "class_sections_json": [],
            "teachers_json": [],
            "subjects_json": [],
            "rooms_json": [],
            "periods_json": [],
            "lesson_allocations_json": [],
            "teacher_availability_json": [],
        })

    academic_years = AcademicYear.objects.select_related("school").filter(school=current_school).order_by("-id")
    selected_academic_year_id = request.GET.get("academic_year_id")
    selected_timetable_id = request.GET.get("timetable_id")

    timetables = Timetable.objects.select_related(
        "school",
        "academic_year"
    ).filter(is_active=True, school=current_school)

    if selected_academic_year_id:
        timetables = timetables.filter(academic_year_id=selected_academic_year_id)

    timetables = timetables.order_by("-id")

    selected_timetable = None

    if selected_timetable_id:
        selected_timetable = timetables.filter(id=selected_timetable_id).first()

    if not selected_timetable:
        selected_timetable = timetables.first()

    if selected_timetable:
        selected_academic_year_id = str(selected_timetable.academic_year_id)
        readiness = timetable_readiness(selected_timetable)
        if not readiness["can_open_builder"]:
            messages.warning(
                request,
                f"Complete timetable setup before opening the builder. Missing: {readiness['missing_text']}."
            )
            return redirect("timetable_config", pk=selected_timetable.pk)

    selected_school = current_school
    configured_section_ids = set()
    configured_teacher_ids = set()
    configured_room_ids = set()
    if selected_timetable:
        configuration = TimetableConfiguration.objects.filter(timetable=selected_timetable).first()
        if configuration:
            configured_section_ids = set(configuration.class_sections.values_list("id", flat=True))
            configured_teacher_ids = set(configuration.teachers.values_list("id", flat=True))
            configured_room_ids = set(configuration.rooms.values_list("id", flat=True))

    class_sections = ClassSection.objects.select_related(
        "school",
        "class_level",
        "division",
        "class_teacher",
        "default_room"
    ).filter(is_active=True)

    if selected_school:
        class_sections = class_sections.filter(school=selected_school)

    if selected_timetable:
        class_sections = class_sections.filter(timetable=selected_timetable)

    if configured_section_ids:
        class_sections = class_sections.filter(id__in=configured_section_ids)

    class_sections = class_sections.order_by(
        "school__name",
        "class_level__sort_order",
        "division__sort_order"
    )

    teachers = Teacher.objects.select_related("school").filter(is_active=True)
    subjects = Subject.objects.select_related("school").filter(is_active=True)
    rooms = Room.objects.select_related("school").filter(is_active=True)

    if selected_school:
        teachers = teachers.filter(school=selected_school)
        subjects = subjects.filter(school=selected_school)
        rooms = rooms.filter(school=selected_school)

    if selected_timetable:
        teachers = teachers.filter(timetable=selected_timetable)
        subjects = subjects.filter(timetable=selected_timetable)
        rooms = rooms.filter(timetable=selected_timetable)

    if configured_teacher_ids:
        teachers = teachers.filter(id__in=configured_teacher_ids)

    if configured_room_ids:
        rooms = rooms.filter(id__in=configured_room_ids)

    teachers = teachers.order_by("name")
    subjects = subjects.order_by("name")
    rooms = rooms.order_by("name")

    periods_data = _builder_periods_data(selected_timetable)

    lesson_allocations = LessonAllocation.objects.select_related(
        "school",
        "timetable",
        "academic_year",
        "class_section",
        "subject",
        "teacher",
        "default_room"
    ).filter(is_active=True)

    if selected_timetable:
        lesson_allocations = lesson_allocations.filter(timetable=selected_timetable)
    elif selected_academic_year_id:
        lesson_allocations = lesson_allocations.filter(academic_year_id=selected_academic_year_id)

    if selected_school:
        lesson_allocations = lesson_allocations.filter(school=selected_school)

    if configured_section_ids:
        lesson_allocations = lesson_allocations.filter(class_section_id__in=configured_section_ids)

    if configured_teacher_ids:
        lesson_allocations = lesson_allocations.filter(teacher_id__in=configured_teacher_ids)

    if configured_room_ids:
        lesson_allocations = lesson_allocations.filter(Q(default_room_id__in=configured_room_ids) | Q(default_room__isnull=True))

    class_sections_data = []

    for section in class_sections:
        class_sections_data.append({
            "id": section.id,
            "name": f"{section.class_level.name} {section.division.name}",
            "class_level": section.class_level.name,
            "division": section.division.name,
            "section_type": section.class_level.section_type,
            "school_id": section.school.id,
            "school_name": section.school.name,
            "class_teacher_id": section.class_teacher.id if section.class_teacher else None,
            "class_teacher_name": section.class_teacher.name if section.class_teacher else "",
            "default_room_id": section.default_room.id if section.default_room else None,
            "default_room_name": section.default_room.name if section.default_room else "",
        })

    teachers_data = []

    for teacher in teachers:
        teachers_data.append({
            "id": teacher.id,
            "name": teacher.name,
            "school_id": teacher.school.id,
            "department": teacher.department,
            "max_day": teacher.max_periods_per_day,
            "max_week": teacher.max_periods_per_week,
        })

    subjects_data = []

    for subject in subjects:
        subjects_data.append({
            "id": subject.id,
            "name": subject.name,
            "short_name": subject.short_name,
            "color": subject.color_code,
            "section_type": subject.section_type,
            "subject_type": subject.subject_type,
            "school_id": subject.school.id,
        })

    rooms_data = []

    for room in rooms:
        rooms_data.append({
            "id": room.id,
            "name": room.name,
            "room_type": room.room_type,
            "school_id": room.school.id,
        })

    lesson_allocations_data = []

    for allocation in lesson_allocations:
        lesson_allocations_data.append({
            "id": allocation.id,
            "class_section_id": allocation.class_section.id,
            "subject_id": allocation.subject.id,
            "subject_name": allocation.subject.name,
            "teacher_id": allocation.teacher.id,
            "teacher_name": allocation.teacher.name,
            "room_id": allocation.default_room.id if allocation.default_room else None,
            "room_name": allocation.default_room.name if allocation.default_room else "",
            "weekly_periods": allocation.weekly_periods,
            "requires_double_period": allocation.requires_double_period,
        })

    teacher_availability_data = []

    availability_query = TeacherAvailability.objects.select_related("teacher", "day", "period").filter(teacher__school=current_school)
    if selected_timetable:
        availability_query = availability_query.filter(teacher__timetable=selected_timetable)
    if configured_teacher_ids:
        availability_query = availability_query.filter(teacher_id__in=configured_teacher_ids)

    for availability in availability_query:
        teacher_availability_data.append({
            "teacher_id": availability.teacher_id,
            "day_id": availability.day_id,
            "period_id": availability.period_id,
            "is_available": availability.is_available,
        })

    context = {
        "academic_years": academic_years,
        "timetables": timetables,
        "selected_academic_year_id": selected_academic_year_id or "",
        "selected_timetable_id": selected_timetable.id if selected_timetable else "",

        "class_sections_json": class_sections_data,
        "teachers_json": teachers_data,
        "subjects_json": subjects_data,
        "rooms_json": rooms_data,
        "periods_json": periods_data,
        "lesson_allocations_json": lesson_allocations_data,
        "teacher_availability_json": teacher_availability_data,
    }
    return render(request, template_name, context)


@login_required
@log_exceptions
def timetable_builder_template_2(request):
    return timetable_builder(request, "timetable_builder_template_2.html")


@login_required
@log_exceptions
def timetable_builder_template_3(request):
    return timetable_builder(request, "timetable_builder_template_3.html")








from .models import  TimetableEntry


@log_exceptions
def _dedupe_by(items, key_func):
    deduped = []
    seen = set()

    for item in items:
        key = key_func(item)

        if key in seen:
            continue

        seen.add(key)
        deduped.append(item)

    return deduped


@log_exceptions
def _format_period_time(value):
    return value.strftime("%H:%M") if value else ""


@log_exceptions
def _builder_bell_schedule(timetable=None, school=None):
    if timetable:
        configuration = TimetableConfiguration.objects.filter(timetable=timetable).select_related("bell_schedule").first()
        if configuration and configuration.bell_schedule_id:
            return configuration.bell_schedule

        bell_schedule = BellSchedule.objects.filter(
            school=timetable.school,
            academic_year=timetable.academic_year,
            timetable=timetable,
            is_active=True
        ).order_by("-id").first()

        if bell_schedule:
            return bell_schedule

    bell_schedules = BellSchedule.objects.filter(is_active=True)
    if school:
        bell_schedules = bell_schedules.filter(school=school)

    return bell_schedules.order_by("-id").first()


@log_exceptions
def _builder_periods_data(timetable=None, school=None):
    bell_schedule = _builder_bell_schedule(timetable, school)
    school = timetable.school if timetable else school or bell_schedule.school if bell_schedule else None
    configured_day_ids = set()
    configured_period_ids = set()

    if timetable:
        configuration = TimetableConfiguration.objects.filter(timetable=timetable).first()
        if configuration:
            configured_day_ids = set(configuration.working_days.values_list("id", flat=True))
            configured_period_ids = set(configuration.periods.values_list("id", flat=True))

    days_query = Day.objects.filter(is_working=True)

    if school:
        days_query = days_query.filter(school=school)
    if timetable:
        days_query = days_query.filter(timetable=timetable)

    if configured_day_ids:
        days_query = days_query.filter(id__in=configured_day_ids)

    days = _dedupe_by(
        days_query.order_by("sort_order", "id"),
        lambda day: (day.name, day.day_type, day.sort_order)
    )

    periods_query = Period.objects.filter(school=school) if school else Period.objects.none()
    if timetable:
        periods_query = periods_query.filter(timetable=timetable)

    if bell_schedule:
        periods_query = periods_query.filter(bell_schedule=bell_schedule)

    if configured_period_ids:
        periods_query = periods_query.filter(id__in=configured_period_ids)

    weekday_periods = periods_query.filter(day_type="WEEKDAY").order_by("period_number", "id")
    saturday_periods = periods_query.filter(day_type="SATURDAY").order_by("period_number", "id")

    weekday_periods = _dedupe_by(
        weekday_periods,
        lambda period: (
            period.day_type,
            period.period_number,
            period.name,
            period.period_type,
            period.start_time,
            period.end_time,
        )
    )

    saturday_periods = _dedupe_by(
        saturday_periods,
        lambda period: (
            period.day_type,
            period.period_number,
            period.name,
            period.period_type,
            period.start_time,
            period.end_time,
        )
    )

    periods_data = []

    if days and (weekday_periods or saturday_periods):
        for day in days:
            selected_periods = saturday_periods if day.day_type == "SATURDAY" else weekday_periods

            for period in selected_periods:
                periods_data.append({
                    "day_id": day.id,
                    "day_name": day.name,
                    "day_type": day.day_type,
                    "period_id": period.id,
                    "period_name": period.name,
                    "period_number": period.period_number,
                    "period_type": period.period_type,
                    "is_teaching_period": period.is_teaching_period,
                    "start_time": _format_period_time(period.start_time),
                    "end_time": _format_period_time(period.end_time),
                })

    return periods_data


@log_exceptions
def _builder_class_sections_data(school, timetable=None):
    class_sections = ClassSection.objects.select_related(
        "school",
        "class_level",
        "division",
        "class_teacher",
        "default_room"
    ).filter(
        school=school,
        is_active=True
    ).order_by(
        "school__name",
        "class_level__sort_order",
        "division__sort_order"
    )
    if timetable:
        configuration = TimetableConfiguration.objects.filter(timetable=timetable).first()
        if configuration:
            configured_section_ids = set(configuration.class_sections.values_list("id", flat=True))
            if configured_section_ids:
                class_sections = class_sections.filter(id__in=configured_section_ids)

    return [{
        "id": section.id,
        "name": f"{section.class_level.name} {section.division.name}",
        "class_level": section.class_level.name,
        "division": section.division.name,
        "section_type": section.class_level.section_type,
        "school_id": section.school.id,
        "school_name": section.school.name,
        "class_teacher_id": section.class_teacher.id if section.class_teacher else None,
        "class_teacher_name": section.class_teacher.name if section.class_teacher else "",
        "default_room_id": section.default_room.id if section.default_room else None,
        "default_room_name": section.default_room.name if section.default_room else "",
    } for section in class_sections]


@log_exceptions
def _builder_subjects_data(school, timetable=None):
    subjects = Subject.objects.select_related("school").filter(school=school, is_active=True)
    if timetable:
        subjects = subjects.filter(timetable=timetable)
    subjects = subjects.order_by("name")

    return [{
        "id": subject.id,
        "name": subject.name,
        "short_name": subject.short_name,
        "color": subject.color_code,
        "section_type": subject.section_type,
        "subject_type": subject.subject_type,
        "school_id": subject.school.id,
    } for subject in subjects]


@log_exceptions
def _builder_rooms_data(school, timetable=None):
    rooms = Room.objects.select_related("school").filter(school=school, is_active=True).order_by("name")
    if timetable:
        rooms = rooms.filter(timetable=timetable)
        configuration = TimetableConfiguration.objects.filter(timetable=timetable).first()
        if configuration:
            configured_room_ids = set(configuration.rooms.values_list("id", flat=True))
            if configured_room_ids:
                rooms = rooms.filter(id__in=configured_room_ids)

    return [{
        "id": room.id,
        "name": room.name,
        "room_type": room.room_type,
        "school_id": room.school.id,
    } for room in rooms]


@log_exceptions
def _entry_data_for_teacher(timetable, teacher):
    data = {}
    occupied = {}

    if not timetable:
        return data, occupied

    entries = TimetableEntry.objects.filter(
        timetable=timetable
    ).select_related(
        "class_section",
        "teacher",
        "subject",
        "room"
    )

    for entry in entries:
        key = f"{entry.class_section_id}_{entry.day_id_value}_{entry.period_id_value}"
        payload = {
            "class_section_id": entry.class_section.id if entry.class_section else None,
            "class_section_name": str(entry.class_section) if entry.class_section else "",
            "teacher_id": entry.teacher.id if entry.teacher else None,
            "teacher_name": entry.teacher.name if entry.teacher else "",
            "subject_id": entry.subject.id if entry.subject else None,
            "subject_name": entry.subject.name if entry.subject else "",
            "room_id": entry.room.id if entry.room else None,
            "room_name": entry.room.name if entry.room else "",
            "is_locked": entry.is_locked,
        }

        if entry.teacher_id == teacher.id:
            data[key] = payload
        else:
            occupied[key] = payload

    return data, occupied


@login_required
@log_exceptions
def teacher_timetable_builder(request, teacher_id):
    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    teacher = get_object_or_404(Teacher, id=teacher_id, school=current_school, is_active=True)
    timetable_id = request.GET.get("timetable_id")

    timetables = Timetable.objects.select_related(
        "school",
        "academic_year"
    ).filter(school=current_school, is_active=True).order_by("-id")

    timetable = None
    if timetable_id:
        timetable = get_object_or_404(Timetable, id=timetable_id, school=current_school)
    else:
        timetable = timetables.first()

    if timetable:
        readiness = timetable_readiness(timetable)
        if not readiness["can_open_builder"]:
            messages.warning(
                request,
                f"Complete timetable setup before opening the teacher builder. Missing: {readiness['missing_text']}."
            )
            return redirect("timetable_config", pk=timetable.pk)
        configuration = TimetableConfiguration.objects.filter(timetable=timetable).first()
        if configuration and configuration.teachers.exists() and not configuration.teachers.filter(pk=teacher.pk).exists():
            messages.warning(request, f"{teacher.name} is not included in this timetable configuration.")
            return redirect("timetable_config", pk=timetable.pk)

    timetable_entries, occupied_entries = _entry_data_for_teacher(timetable, teacher)

    context = {
        "teacher": teacher,
        "timetable": timetable,
        "timetables": timetables,
        "class_sections_json": _builder_class_sections_data(current_school, timetable),
        "subjects_json": _builder_subjects_data(current_school, timetable),
        "rooms_json": _builder_rooms_data(current_school, timetable),
        "periods_json": _builder_periods_data(timetable, current_school),
        "timetable_entries_json": timetable_entries,
        "occupied_entries_json": occupied_entries,
    }

    return render(request, "teacher_timetable_builder.html", context)


@csrf_exempt
@login_required
@log_exceptions
def validate_timetable_entries(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request"})

    data = json.loads(request.body)
    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    timetable_id = data.get("timetable_id")
    entries = data.get("entries", [])

    if not timetable_id:
        return JsonResponse({"success": False, "message": "Timetable is required"})

    timetable = get_object_or_404(Timetable, id=timetable_id, school=current_school)
    readiness = timetable_readiness(timetable)
    if not readiness["can_open_builder"]:
        return JsonResponse({
            "success": False,
            "message": f"Complete timetable setup before validating. Missing: {readiness['missing_text']}.",
            "validation": {
                "errors": [f"Setup pending: {readiness['missing_text']}."],
                "warnings": [],
                "summary": {},
            },
        })

    validation = _validate_timetable_payload(current_school, timetable, entries)

    return JsonResponse({
        "success": not validation["errors"],
        "message": _validation_message(validation, saving=False),
        "validation": validation,
    })


@log_exceptions
def _validation_message(validation, saving=False):
    error_count = len(validation["errors"])
    warning_count = len(validation["warnings"])

    if error_count:
        return f"Timetable has {error_count} critical issue(s). Fix them before saving."

    if warning_count:
        action = "saved" if saving else "valid"
        return f"Timetable is {action} with {warning_count} warning(s). Review the audit details."

    return "Timetable validation passed." if not saving else "Timetable saved successfully."


@log_exceptions
def _room_type_warning(subject, room):
    if not subject or not room:
        return ""

    if subject.subject_type == "PRACTICAL" and room.room_type == "CLASSROOM":
        return (
            f"{subject.name} is a practical subject, so it should use a lab or activity room "
            f"instead of classroom '{room.name}'."
        )

    return ""


@log_exceptions
def _validate_timetable_payload(current_school, timetable, entries):
    errors = []
    warnings = []
    summary = {
        "entries": len(entries),
        "teacher_conflicts": 0,
        "room_conflicts": 0,
        "missing_allocations": 0,
        "extra_allocations": 0,
        "teacher_unavailable": 0,
        "teacher_overloads": 0,
        "class_without_room": 0,
        "capability_mismatches": 0,
        "class_teacher_issues": 0,
    }

    periods_by_key = {
        (str(item["day_id"]), str(item["period_id"])): item
        for item in _builder_periods_data(timetable=timetable)
    }
    configured_section_ids = set()
    configured_teacher_ids = set()
    configured_room_ids = set()
    configuration = TimetableConfiguration.objects.filter(timetable=timetable).first()
    if configuration:
        configured_section_ids = set(configuration.class_sections.values_list("id", flat=True))
        configured_teacher_ids = set(configuration.teachers.values_list("id", flat=True))
        configured_room_ids = set(configuration.rooms.values_list("id", flat=True))

    if not configured_section_ids:
        configured_section_ids = set(ClassSection.objects.filter(timetable=timetable, is_active=True).values_list("id", flat=True))
    if not configured_teacher_ids:
        configured_teacher_ids = set(Teacher.objects.filter(timetable=timetable, is_active=True).values_list("id", flat=True))
    if not configured_room_ids:
        configured_room_ids = set(Room.objects.filter(timetable=timetable, is_active=True).values_list("id", flat=True))

    class_sections_query = ClassSection.objects.select_related("class_level", "division", "class_teacher").filter(school=current_school, timetable=timetable, is_active=True)
    if configured_section_ids:
        class_sections_query = class_sections_query.filter(id__in=configured_section_ids)
    class_sections = {
        str(section.id): section
        for section in class_sections_query
    }

    teachers_query = Teacher.objects.filter(school=current_school, timetable=timetable, is_active=True)
    if configured_teacher_ids:
        teachers_query = teachers_query.filter(id__in=configured_teacher_ids)
    teachers = {
        str(teacher.id): teacher
        for teacher in teachers_query
    }
    subjects = {
        str(subject.id): subject
        for subject in Subject.objects.filter(school=current_school, timetable=timetable, is_active=True)
    }
    rooms_query = Room.objects.filter(school=current_school, timetable=timetable, is_active=True)
    if configured_room_ids:
        rooms_query = rooms_query.filter(id__in=configured_room_ids)
    rooms = {
        str(room.id): room
        for room in rooms_query
    }
    capability_exact_sections = set()
    capability_class_levels = set()
    capability_broad = set()
    capabilities = TeacherSubjectCapability.objects.filter(
        school=current_school,
        timetable=timetable,
    ).prefetch_related("class_sections", "class_levels")

    for capability in capabilities:
        pair = (str(capability.teacher_id), str(capability.subject_id))
        section_ids = {str(section_id) for section_id in capability.class_sections.values_list("id", flat=True)}
        level_ids = {str(level_id) for level_id in capability.class_levels.values_list("id", flat=True)}

        if not section_ids and not level_ids:
            capability_broad.add(pair)

        for section_id in section_ids:
            capability_exact_sections.add((*pair, section_id))

        for level_id in level_ids:
            capability_class_levels.add((*pair, level_id))

    teacher_slots = defaultdict(list)
    room_slots = defaultdict(list)
    class_slots = defaultdict(list)
    teacher_day_load = defaultdict(int)
    teacher_week_load = defaultdict(int)
    allocation_load = defaultdict(int)

    for index, entry in enumerate(entries, start=1):
        class_id = str(entry.get("class_section_id") or "")
        day_id = str(entry.get("day_id") or "")
        period_id = str(entry.get("period_id") or "")
        teacher_id = str(entry.get("teacher_id") or "")
        subject_id = str(entry.get("subject_id") or "")
        room_id = str(entry.get("room_id") or "")
        label = f"Row {index}: {entry.get('day_name', '')} {entry.get('period_name', '')}".strip()

        if class_id not in class_sections:
            errors.append(f"{label}: invalid or inactive class section.")
            continue

        period = periods_by_key.get((day_id, period_id))
        if not period:
            errors.append(f"{label}: invalid period for this timetable bell schedule.")
            continue

        if not period.get("is_teaching_period"):
            errors.append(f"{label}: lecture assigned in non-teaching period '{period.get('period_name')}'.")

        class_slot_key = (class_id, day_id, period_id)
        class_slots[class_slot_key].append(label)

        if subject_id and subject_id not in subjects:
            errors.append(f"{label}: invalid or inactive subject.")

        if teacher_id and teacher_id not in teachers:
            errors.append(f"{label}: invalid or inactive teacher.")

        if room_id and room_id not in rooms:
            errors.append(f"{label}: invalid or inactive room.")

        if not room_id:
            summary["class_without_room"] += 1
            class_name = str(class_sections.get(class_id, "Class"))
            subject_name = subjects.get(subject_id).name if subject_id in subjects else "Subject"
            warnings.append(f"No room assigned: {class_name} - {entry.get('day_name', '')} {entry.get('period_name', '')} - {subject_name}.")

        if teacher_id and teacher_id in teachers:
            teacher_slots[(teacher_id, day_id, period_id)].append(label)
            teacher_day_load[(teacher_id, day_id)] += 1
            teacher_week_load[teacher_id] += 1

            unavailable = TeacherAvailability.objects.filter(
                teacher_id=teacher_id,
                day_id=day_id,
                period_id=period_id,
                is_available=False,
            ).exists()
            if unavailable:
                summary["teacher_unavailable"] += 1
                errors.append(f"{label}: {teachers[teacher_id].name} is unavailable in this period.")

        if class_id in class_sections and teacher_id in teachers and subject_id in subjects:
            class_section = class_sections[class_id]
            pair = (teacher_id, subject_id)
            has_capability = (
                pair in capability_broad or
                (*pair, class_id) in capability_exact_sections or
                (*pair, str(class_section.class_level_id)) in capability_class_levels
            )
            if not has_capability:
                summary["capability_mismatches"] += 1
                errors.append(
                    f"{label}: {teachers[teacher_id].name} does not have Teacher Subject Capability for {class_section} - {subjects[subject_id].name}."
                )

        if room_id and room_id in rooms:
            if subject_id in subjects:
                room_type_warning = _room_type_warning(subjects[subject_id], rooms[room_id])
                if room_type_warning:
                    warnings.append(f"{label}: {room_type_warning}")
            room_slots[(room_id, day_id, period_id)].append(label)

        if class_id and subject_id and teacher_id:
            allocation_load[(class_id, subject_id, teacher_id)] += 1

    for labels in class_slots.values():
        if len(labels) > 1:
            errors.append(f"Class slot duplicate: {', '.join(labels)}.")

    for (teacher_id, day_id, period_id), labels in teacher_slots.items():
        if len(labels) > 1:
            summary["teacher_conflicts"] += 1
            teacher_name = teachers.get(teacher_id).name if teacher_id in teachers else "Teacher"
            errors.append(f"{teacher_name} is double-booked in the same period: {', '.join(labels)}.")

    for (room_id, day_id, period_id), labels in room_slots.items():
        if len(labels) > 1:
            summary["room_conflicts"] += 1
            room_name = rooms.get(room_id).name if room_id in rooms else "Room"
            errors.append(f"{room_name} is double-booked in the same period: {', '.join(labels)}.")

    teacher_required_week = defaultdict(int)
    class_teacher_allocation_load = defaultdict(int)
    allocations = LessonAllocation.objects.select_related("class_section", "subject", "teacher").filter(
        school=current_school,
        timetable=timetable,
        is_active=True,
    )
    if configured_section_ids:
        allocations = allocations.filter(class_section_id__in=configured_section_ids)
    if configured_teacher_ids:
        allocations = allocations.filter(teacher_id__in=configured_teacher_ids)
    if configured_room_ids:
        allocations = allocations.filter(Q(default_room_id__in=configured_room_ids) | Q(default_room__isnull=True))
    allocation_keys = set()
    for allocation in allocations:
        key = (str(allocation.class_section_id), str(allocation.subject_id), str(allocation.teacher_id))
        allocation_keys.add(key)
        teacher_required_week[str(allocation.teacher_id)] += allocation.weekly_periods
        class_teacher_allocation_load[(str(allocation.class_section_id), str(allocation.teacher_id))] += allocation.weekly_periods
        placed = allocation_load.get(key, 0)
        if placed < allocation.weekly_periods:
            summary["missing_allocations"] += allocation.weekly_periods - placed
            warnings.append(
                f"Missing {allocation.weekly_periods - placed} period(s): {allocation.teacher.name} - {allocation.class_section} - {allocation.subject.name}."
            )
        elif placed > allocation.weekly_periods:
            summary["extra_allocations"] += placed - allocation.weekly_periods
            warnings.append(
                f"Extra {placed - allocation.weekly_periods} period(s): {allocation.teacher.name} - {allocation.class_section} - {allocation.subject.name}."
            )

    for key, placed in allocation_load.items():
        if key not in allocation_keys:
            class_id, subject_id, teacher_id = key
            warnings.append(
                f"No active lesson allocation found for {teachers.get(teacher_id, 'Teacher')} - {class_sections.get(class_id, 'Class')} - {subjects.get(subject_id, 'Subject')}."
            )

    first_teaching_periods_by_day = {}
    for period in periods_by_key.values():
        if not period.get("is_teaching_period"):
            continue
        day_id = str(period["day_id"])
        existing = first_teaching_periods_by_day.get(day_id)
        if not existing or period.get("period_number", 0) < existing.get("period_number", 0):
            first_teaching_periods_by_day[day_id] = period

    first_lecture_required_count = len(first_teaching_periods_by_day)
    if first_lecture_required_count:
        for class_section in class_sections.values():
            if not class_section.class_teacher_id:
                continue

            class_teacher_id = str(class_section.class_teacher_id)
            allocated_count = class_teacher_allocation_load.get((str(class_section.id), class_teacher_id), 0)
            teacher_name = class_section.class_teacher.name if class_section.class_teacher else "Class teacher"

            if not allocated_count:
                summary["class_teacher_issues"] += 1
                warnings.append(
                    f"Class teacher priority not ready: {teacher_name} is class teacher for {class_section}, but has no active lesson allocation for this class."
                )
            elif allocated_count < first_lecture_required_count:
                summary["class_teacher_issues"] += 1
                warnings.append(
                    f"Class teacher priority partial: {teacher_name} has {allocated_count} allocated period(s) for {class_section}, but {first_lecture_required_count} first-lecture slot(s) are needed for all working days."
                )

    for teacher_id, load in teacher_week_load.items():
        teacher = teachers.get(teacher_id)
        if not teacher:
            continue
        effective_week_limit = max(teacher.max_periods_per_week or 0, teacher_required_week.get(teacher_id, 0))
        if effective_week_limit and load > effective_week_limit:
            summary["teacher_overloads"] += 1
            errors.append(f"{teacher.name} has {load} periods/week, above allowed {effective_week_limit}.")

    teaching_day_count = len({item["day_id"] for item in periods_by_key.values() if item.get("is_teaching_period")}) or 1
    for (teacher_id, day_id), load in teacher_day_load.items():
        teacher = teachers.get(teacher_id)
        if not teacher:
            continue
        effective_day_limit = max(
            teacher.max_periods_per_day or 0,
            (teacher_required_week.get(teacher_id, 0) + teaching_day_count - 1) // teaching_day_count,
        )
        if effective_day_limit and load > effective_day_limit:
            summary["teacher_overloads"] += 1
            errors.append(f"{teacher.name} has {load} periods on one day, above allowed {effective_day_limit}.")

    return {
        "errors": errors,
        "warnings": warnings,
        "summary": summary,
    }


@log_exceptions
def _teacher_timetable_impact(timetable, teacher, entries):
    target_keys = [
        (
            str(entry["class_section_id"]),
            str(entry["day_id"]),
            str(entry["period_id"]),
        )
        for entry in entries
    ]
    target_rooms = [
        (
            str(entry.get("room_id")),
            str(entry["day_id"]),
            str(entry["period_id"]),
        )
        for entry in entries
        if entry.get("room_id")
    ]

    teacher_entries = TimetableEntry.objects.filter(timetable=timetable, teacher=teacher)
    class_overlaps = TimetableEntry.objects.none()
    room_overlaps = TimetableEntry.objects.none()

    for class_section_id, day_id, period_id in target_keys:
        class_overlaps = class_overlaps | TimetableEntry.objects.filter(
            timetable=timetable,
            class_section_id=class_section_id,
            day_id_value=day_id,
            period_id_value=period_id,
        ).exclude(teacher=teacher)

    for room_id, day_id, period_id in target_rooms:
        room_overlaps = room_overlaps | TimetableEntry.objects.filter(
            timetable=timetable,
            room_id=room_id,
            day_id_value=day_id,
            period_id_value=period_id,
        ).exclude(teacher=teacher)

    impacted_ids = set(teacher_entries.values_list("id", flat=True))
    impacted_ids.update(class_overlaps.values_list("id", flat=True))
    impacted_ids.update(room_overlaps.values_list("id", flat=True))

    locked_count = TimetableEntry.objects.filter(id__in=impacted_ids, is_locked=True).count()

    return {
        "teacher_entries": teacher_entries.count(),
        "class_overlaps": class_overlaps.distinct().count(),
        "room_overlaps": room_overlaps.distinct().count(),
        "total_replaced": len(impacted_ids),
        "locked_entries": locked_count,
    }


@log_exceptions
def _timetable_save_impact(timetable, entries, scope_class_section_ids=None):
    scope_class_section_ids = {str(item) for item in scope_class_section_ids or []}
    incoming_keys = {
        (
            str(entry["class_section_id"]),
            str(entry["day_id"]),
            str(entry["period_id"]),
        )
        for entry in entries
    }
    existing_entries = TimetableEntry.objects.filter(timetable=timetable)
    if scope_class_section_ids:
        existing_entries = existing_entries.filter(class_section_id__in=scope_class_section_ids)

    existing_keys = {
        (
            str(entry.class_section_id),
            str(entry.day_id_value),
            str(entry.period_id_value),
        )
        for entry in existing_entries
    }

    locked_count = existing_entries.filter(is_locked=True).count()
    replace_count = len(existing_keys & incoming_keys)
    delete_count = len(existing_keys - incoming_keys)
    create_count = len(incoming_keys - existing_keys)

    return {
        "scope": "section_scope" if scope_class_section_ids else "full_timetable",
        "scope_sections": len(scope_class_section_ids),
        "incoming_entries": len(incoming_keys),
        "existing_entries": len(existing_keys),
        "entries_to_create": create_count,
        "entries_to_replace": replace_count,
        "entries_to_delete": delete_count,
        "locked_entries": locked_count,
        "total_affected": len(existing_keys),
    }


@csrf_exempt
@login_required
@log_exceptions
def save_timetable_entries(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request"})

    data = json.loads(request.body)
    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    timetable_id = data.get("timetable_id")
    entries = data.get("entries", [])
    preview_only = data.get("preview_only")
    confirmed_impact = data.get("confirmed_impact")
    allow_empty_save = data.get("allow_empty_save")
    save_scope = data.get("save_scope") or "full_timetable"
    scope_class_section_ids = {str(item) for item in data.get("scope_class_section_ids", []) if item}

    if not timetable_id:
        return JsonResponse({"success": False, "message": "Timetable is required"})

    timetable = get_object_or_404(Timetable, id=timetable_id, school=current_school)
    readiness = timetable_readiness(timetable)
    if not readiness["can_open_builder"]:
        return JsonResponse({
            "success": False,
            "message": f"Complete timetable setup before saving. Missing: {readiness['missing_text']}.",
            "validation": {
                "errors": [f"Setup pending: {readiness['missing_text']}."],
                "warnings": [],
                "summary": {},
            },
        })

    if save_scope == "section_scope":
        if not scope_class_section_ids:
            return JsonResponse({"success": False, "message": "Select at least one visible section before scoped save."})

        configured_section_ids = set()
        configuration = TimetableConfiguration.objects.filter(timetable=timetable).first()
        if configuration:
            configured_section_ids = {str(item) for item in configuration.class_sections.values_list("id", flat=True)}

        if configured_section_ids and not scope_class_section_ids.issubset(configured_section_ids):
            return JsonResponse({"success": False, "message": "Scoped save contains sections outside this timetable configuration."})

        valid_section_ids = {
            str(item) for item in ClassSection.objects.filter(
                id__in=scope_class_section_ids,
                school=current_school,
                is_active=True,
            ).values_list("id", flat=True)
        }
        if valid_section_ids != scope_class_section_ids:
            return JsonResponse({"success": False, "message": "Scoped save contains invalid or inactive sections."})

        entries = [
            entry for entry in entries
            if str(entry.get("class_section_id") or "") in scope_class_section_ids
        ]
    else:
        scope_class_section_ids = set()

    validation = _validate_timetable_payload(current_school, timetable, entries)
    impact = _timetable_save_impact(timetable, entries, scope_class_section_ids)
    if not entries and impact["existing_entries"] and not allow_empty_save:
        return JsonResponse({
            "success": False,
            "message": "Empty timetable save blocked to protect existing saved slots. Remove slots manually only after rebuilding the board.",
            "validation": {
                "errors": ["Empty save blocked because this timetable already has saved slots."],
                "warnings": [],
                "summary": {"entries": 0},
            },
            "impact": impact,
        })

    if validation["errors"]:
        return JsonResponse({
            "success": False,
            "message": _validation_message(validation, saving=True),
            "validation": validation,
            "impact": impact,
        })

    if preview_only:
        return JsonResponse({
            "success": True,
            "message": "Timetable save impact preview ready.",
            "validation": validation,
            "impact": impact,
        })

    if impact["total_affected"] and not confirmed_impact:
        return JsonResponse({
            "success": False,
            "requires_confirmation": True,
            "message": "This save will replace existing timetable slots. Please confirm the impact preview first.",
            "validation": validation,
            "impact": impact,
        })

    try:
        with transaction.atomic():
            entries_to_delete = TimetableEntry.objects.filter(timetable=timetable)
            if scope_class_section_ids:
                entries_to_delete = entries_to_delete.filter(class_section_id__in=scope_class_section_ids)
            entries_to_delete.delete()

            for entry in entries:
                class_section = get_object_or_404(ClassSection, id=entry["class_section_id"], school=current_school, timetable=timetable)

                subject = Subject.objects.filter(id=entry.get("subject_id"), school=current_school, timetable=timetable).first()
                teacher = Teacher.objects.filter(id=entry.get("teacher_id"), school=current_school, timetable=timetable).first()
                room = Room.objects.filter(id=entry.get("room_id"), school=current_school, timetable=timetable).first()

                TimetableEntry.objects.create(
                    timetable=timetable,
                    class_section=class_section,
                    day_id_value=entry["day_id"],
                    day_name=entry["day_name"],
                    period_id_value=entry["period_id"],
                    period_name=entry["period_name"],
                    subject=subject,
                    teacher=teacher,
                    room=room,
                    is_locked=entry.get("is_locked", False),
                )
    except IntegrityError:
        return JsonResponse({
            "success": False,
            "message": "Timetable could not be saved because a duplicate teacher, room, or class slot was detected. Please run Validate and fix the highlighted issue.",
            "validation": validation,
        })

    return JsonResponse({
        "success": True,
        "message": _validation_message(validation, saving=True),
        "validation": validation,
        "impact": impact,
    })


@csrf_exempt
@login_required
@log_exceptions
def save_teacher_timetable_entries(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request"})

    data = json.loads(request.body)
    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    timetable_id = data.get("timetable_id")
    teacher_id = data.get("teacher_id")
    entries = data.get("entries", [])

    if not timetable_id or not teacher_id:
        return JsonResponse({"success": False, "message": "Timetable and teacher are required"})

    timetable = get_object_or_404(Timetable, id=timetable_id, school=current_school)
    teacher = get_object_or_404(Teacher, id=teacher_id, school=current_school)
    readiness = timetable_readiness(timetable)
    if not readiness["can_open_builder"]:
        return JsonResponse({
            "success": False,
            "message": f"Complete timetable setup before saving. Missing: {readiness['missing_text']}.",
            "validation": {
                "errors": [f"Setup pending: {readiness['missing_text']}."],
                "warnings": [],
                "summary": {},
            },
        })

    target_keys = [
        (
            str(entry["class_section_id"]),
            str(entry["day_id"]),
            str(entry["period_id"]),
        )
        for entry in entries
    ]
    teacher_period_keys = [(day_id, period_id) for _, day_id, period_id in target_keys]

    if len(teacher_period_keys) != len(set(teacher_period_keys)):
        return JsonResponse({
            "success": False,
            "message": "This teacher cannot be assigned to more than one class in the same period."
        })

    impact = _teacher_timetable_impact(timetable, teacher, entries)

    validation_entries = []
    for entry in entries:
        validation_entries.append({
            **entry,
            "teacher_id": teacher.id,
            "teacher_name": teacher.name,
        })

    validation = _validate_timetable_payload(current_school, timetable, validation_entries)
    if validation["errors"]:
        return JsonResponse({
            "success": False,
            "message": _validation_message(validation, saving=True),
            "validation": validation,
            "impact": impact,
        })

    if data.get("preview_only"):
        return JsonResponse({
            "success": True,
            "message": "Teacher timetable impact preview ready.",
            "impact": impact,
            "validation": validation,
        })

    if impact["total_replaced"] and not data.get("confirmed_impact"):
        return JsonResponse({
            "success": False,
            "requires_confirmation": True,
            "message": "This save will replace existing master timetable entries. Please confirm the impact preview first.",
            "impact": impact,
            "validation": validation,
        })

    with transaction.atomic():
        TimetableEntry.objects.filter(timetable=timetable, teacher=teacher).delete()

        for class_section_id, day_id, period_id in target_keys:
            TimetableEntry.objects.filter(
                timetable=timetable,
                class_section_id=class_section_id,
                day_id_value=day_id,
                period_id_value=period_id,
            ).exclude(teacher=teacher).delete()

        for entry in entries:
            room_id = entry.get("room_id")

            if not room_id:
                continue

            TimetableEntry.objects.filter(
                timetable=timetable,
                room_id=room_id,
                day_id_value=str(entry["day_id"]),
                period_id_value=str(entry["period_id"]),
            ).exclude(teacher=teacher).delete()

        for entry in entries:
            class_section = get_object_or_404(ClassSection, id=entry["class_section_id"], school=current_school, timetable=timetable)

            subject = Subject.objects.filter(id=entry.get("subject_id"), school=current_school, timetable=timetable).first()
            room = Room.objects.filter(id=entry.get("room_id"), school=current_school, timetable=timetable).first()

            TimetableEntry.objects.create(
                timetable=timetable,
                class_section=class_section,
                day_id_value=entry["day_id"],
                day_name=entry["day_name"],
                period_id_value=entry["period_id"],
                period_name=entry["period_name"],
                subject=subject,
                teacher=teacher,
                room=room,
                is_locked=entry.get("is_locked", False),
            )

    return JsonResponse({
        "success": True,
        "message": f"{teacher.name}'s timetable saved successfully",
        "impact": impact,
        "validation": validation,
    })








from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from Academic.models import AcademicYear, BellSchedule
from Timetables.models import Timetable


@csrf_exempt
@login_required
@log_exceptions
def create_timetable_api(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request"})

    data = json.loads(request.body)
    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    name = data.get("name")
    academic_year_id = data.get("academic_year_id")
    timetable_type = data.get("timetable_type", "PRIMARY")

    if not name or not academic_year_id:
        return JsonResponse({
            "success": False,
            "message": "Timetable name and academic year are required."
        })

    academic_year = get_object_or_404(AcademicYear, id=academic_year_id, school=current_school)

    bell_schedule = BellSchedule.objects.filter(
        school=current_school,
        academic_year=academic_year
    ).first()

    if not bell_schedule:
        return JsonResponse({
            "success": False,
            "message": "Please create Bell Schedule first."
        })

    timetable = Timetable.objects.create(
        school=current_school,
        academic_year=academic_year,
        name=name,
        timetable_type=timetable_type,
        is_active=True
    )

    return JsonResponse({
        "success": True,
        "id": timetable.id,
        "name": timetable.name,
        "message": "Timetable created successfully."
    })






@csrf_exempt
@login_required
@log_exceptions
def load_timetable_entries(request):
    timetable_id = request.GET.get("timetable_id")

    if not timetable_id:
        return JsonResponse({"success": False, "entries": {}})

    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session.", "entries": {}}, status=403)

    timetable = get_object_or_404(Timetable, id=timetable_id, school=current_school)

    entries = TimetableEntry.objects.filter(
        timetable=timetable
    ).select_related(
        "class_section",
        "teacher",
        "subject",
        "room"
    )

    data = {}

    for entry in entries:
        key = f"{entry.class_section_id}_{entry.day_id_value}_{entry.period_id_value}"

        data[key] = {
            "teacher_id": entry.teacher.id if entry.teacher else None,
            "teacher_name": entry.teacher.name if entry.teacher else "",
            "subject_id": entry.subject.id if entry.subject else None,
            "subject_name": entry.subject.name if entry.subject else "",
            "room_id": entry.room.id if entry.room else None,
            "room_name": entry.room.name if entry.room else "",
            "is_locked": entry.is_locked,
        }

    return JsonResponse({
        "success": True,
        "entries": data
    })


@log_exceptions
def _parse_adjustment_date(value):
    if not value:
        return timezone.localdate()

    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return timezone.localdate()


@log_exceptions
def _status_data(status):
    return {
        "id": status.id,
        "teacher_id": status.teacher_id,
        "teacher_name": status.teacher.name,
        "status_type": status.status_type,
        "status_label": status.get_status_type_display(),
        "full_day": status.full_day,
        "period_ids": [str(period_id) for period_id in status.unavailable_periods.values_list("id", flat=True)],
        "reason": status.reason,
        "notes": status.notes,
    }


@log_exceptions
def _adjustment_data(adjustment):
    return {
        "id": adjustment.id,
        "entry_id": adjustment.original_entry_id,
        "proxy_teacher_id": adjustment.proxy_teacher_id,
        "proxy_teacher_name": adjustment.proxy_teacher.name if adjustment.proxy_teacher else "",
        "status": adjustment.status,
        "reason": adjustment.reason,
        "admin_note": adjustment.admin_note,
        "is_locked": adjustment.is_locked,
    }


@log_exceptions
def _period_is_status_covered(status, period_id):
    if status.full_day:
        return True

    return str(period_id) in {str(item) for item in status.unavailable_periods.values_list("id", flat=True)}


@log_exceptions
def _teacher_unavailable_status(statuses, teacher_id, period_id):
    for status in statuses:
        if status.teacher_id == teacher_id and _period_is_status_covered(status, period_id):
            return status

    return None


@log_exceptions
def _date_day_name(adjustment_date):
    return adjustment_date.strftime("%A")


@log_exceptions
def _teacher_day_load(timetable, adjustment_date, teacher_id):
    day_name = _date_day_name(adjustment_date)
    master_count = TimetableEntry.objects.filter(
        timetable=timetable,
        teacher_id=teacher_id,
        day_name__iexact=day_name,
    ).count()
    proxy_count = LectureAdjustment.objects.filter(
        timetable=timetable,
        date=adjustment_date,
        proxy_teacher_id=teacher_id,
        status="ASSIGNED",
    ).count()
    return master_count + proxy_count


@log_exceptions
def _teaching_period_count(timetable, day_name=None):
    periods = [period for period in _builder_periods_data(timetable) if period["is_teaching_period"]]

    if day_name:
        periods = [period for period in periods if period["day_name"].lower() == day_name.lower()]

    return len(periods)


@log_exceptions
def _teacher_load_payload(timetable, adjustment_date, teacher):
    day_name = _date_day_name(adjustment_date)
    total = _teaching_period_count(timetable, day_name)
    load = _teacher_day_load(timetable, adjustment_date, teacher.id)

    return {
        "id": teacher.id,
        "name": teacher.name,
        "department": teacher.department,
        "day_load": load,
        "free_periods": max(0, total - load),
        "total_periods": total,
    }


@log_exceptions
def _teacher_has_period_conflict(timetable, adjustment_date, teacher_id, day_id, period_id, original_entry_id=None):
    master_conflict = TimetableEntry.objects.filter(
        timetable=timetable,
        teacher_id=teacher_id,
        day_id_value=str(day_id),
        period_id_value=str(period_id),
    )

    if original_entry_id:
        master_conflict = master_conflict.exclude(id=original_entry_id)

    if master_conflict.exists():
        return True

    return LectureAdjustment.objects.filter(
        timetable=timetable,
        date=adjustment_date,
        proxy_teacher_id=teacher_id,
        period_id_value=str(period_id),
        status="ASSIGNED",
    ).exclude(original_entry_id=original_entry_id).exists()


@log_exceptions
def _teacher_static_available(teacher_id, day_id, period_id):
    availability = TeacherAvailability.objects.filter(
        teacher_id=teacher_id,
        day_id=day_id,
        period_id=period_id,
    ).first()

    return not availability or availability.is_available


@log_exceptions
def _suggest_proxy_teachers(entry, adjustment_date, statuses):
    if not entry.teacher_id:
        return []

    teachers = Teacher.objects.filter(
        school=entry.timetable.school,
        timetable=entry.timetable,
        is_active=True,
    ).order_by("name")
    lesson_allocations = LessonAllocation.objects.filter(
        school=entry.timetable.school,
        timetable=entry.timetable,
        is_active=True,
    )
    class_teacher_id = entry.class_section.class_teacher_id if entry.class_section else None
    suggestions = []

    for teacher in teachers:
        if teacher.id == entry.teacher_id:
            continue

        unavailable = _teacher_unavailable_status(statuses, teacher.id, entry.period_id_value)
        if unavailable:
            continue

        if not _teacher_static_available(teacher.id, entry.day_id_value, entry.period_id_value):
            continue

        if _teacher_has_period_conflict(
            entry.timetable,
            adjustment_date,
            teacher.id,
            entry.day_id_value,
            entry.period_id_value,
            entry.id,
        ):
            continue

        score = 0
        tags = []

        if lesson_allocations.filter(class_section=entry.class_section, teacher=teacher).exists():
            score += 45
            tags.append("teaches this class")

        if entry.subject_id and lesson_allocations.filter(subject=entry.subject, teacher=teacher).exists():
            score += 35
            tags.append("same subject")

        if class_teacher_id and teacher.id == class_teacher_id:
            score += 25
            tags.append("class teacher")

        day_load = _teacher_day_load(entry.timetable, adjustment_date, teacher.id)
        score += max(0, 20 - day_load)
        total_periods = _teaching_period_count(entry.timetable, entry.day_name)

        suggestions.append({
            "id": teacher.id,
            "name": teacher.name,
            "department": teacher.department,
            "day_load": day_load,
            "free_periods": max(0, total_periods - day_load),
            "total_periods": total_periods,
            "score": score,
            "tags": tags or ["free teacher"],
        })

    return sorted(suggestions, key=lambda item: (-item["score"], item["day_load"], item["name"]))[:8]


@log_exceptions
def _entry_adjustment_payload(entry, adjustment_date, statuses, adjustment=None):
    covered_status = _teacher_unavailable_status(statuses, entry.teacher_id, entry.period_id_value)
    effective_adjustment = adjustment

    if not effective_adjustment:
        effective_adjustment = LectureAdjustment.objects.filter(
            date=adjustment_date,
            original_entry=entry,
        ).select_related("proxy_teacher").first()

    return {
        "entry_id": entry.id,
        "class_section": str(entry.class_section),
        "subject": entry.subject.name if entry.subject else "Subject",
        "room": entry.room.name if entry.room else "",
        "original_teacher_id": entry.teacher_id,
        "original_teacher": entry.teacher.name if entry.teacher else "",
        "day_id": entry.day_id_value,
        "day_name": entry.day_name,
        "period_id": entry.period_id_value,
        "period_name": entry.period_name,
        "reason": covered_status.reason if covered_status else "",
        "status_label": covered_status.get_status_type_display() if covered_status else "Manual adjustment",
        "teacher_status_id": covered_status.id if covered_status else None,
        "adjustment": _adjustment_data(effective_adjustment) if effective_adjustment else None,
        "suggestions": _suggest_proxy_teachers(entry, adjustment_date, statuses),
    }


@log_exceptions
def _affected_proxy_entries(timetable, adjustment_date, statuses, manual_teacher_id=None):
    day_name = _date_day_name(adjustment_date)
    entries = TimetableEntry.objects.filter(
        timetable=timetable,
        day_name__iexact=day_name,
    ).select_related("timetable", "class_section", "class_section__class_teacher", "subject", "teacher", "room")

    affected_entries = []
    seen_entry_ids = set()

    for entry in entries:
        covered_status = _teacher_unavailable_status(statuses, entry.teacher_id, entry.period_id_value)
        manual_match = manual_teacher_id and str(entry.teacher_id) == str(manual_teacher_id)
        existing_adjustment = LectureAdjustment.objects.filter(date=adjustment_date, original_entry=entry).first()

        if not covered_status and not manual_match and not existing_adjustment:
            continue

        affected_entries.append((entry, existing_adjustment))
        seen_entry_ids.add(entry.id)

    saved_adjustments = LectureAdjustment.objects.filter(
        timetable=timetable,
        date=adjustment_date,
    ).select_related("original_entry", "class_section", "subject", "room", "original_teacher", "proxy_teacher")

    for adjustment in saved_adjustments:
        if adjustment.original_entry_id in seen_entry_ids:
            continue

        affected_entries.append((adjustment.original_entry, adjustment))

    return affected_entries


@login_required
@log_exceptions
def proxy_adjustment_panel(request):
    current_school = get_current_school(request)
    timetables = Timetable.objects.select_related("school", "academic_year").filter(is_active=True)
    if current_school:
        timetables = timetables.filter(school=current_school)
    timetables = timetables.order_by("-id")

    selected_timetable_id = request.GET.get("timetable_id")
    selected_timetable = timetables.filter(id=selected_timetable_id).first() if selected_timetable_id else None
    if not selected_timetable:
        selected_timetable = (
            timetables.annotate(
                active_teacher_count=Count("teachers", filter=Q(teachers__is_active=True), distinct=True),
                entry_count=Count("entries", distinct=True),
            )
            .filter(active_teacher_count__gt=0)
            .order_by("-entry_count", "-id")
            .first()
            or timetables.first()
        )

    teachers = Teacher.objects.select_related("school").filter(is_active=True)

    if selected_timetable:
        teachers = teachers.filter(school=selected_timetable.school, timetable=selected_timetable)
    elif current_school:
        teachers = teachers.filter(school=current_school)

    context = {
        "timetables": timetables,
        "selected_timetable_id": selected_timetable.id if selected_timetable else "",
        "selected_date": timezone.localdate().strftime("%Y-%m-%d"),
        "teachers_json": [{
            "id": teacher.id,
            "name": teacher.name,
            "school_id": teacher.school_id,
            "department": teacher.department,
        } for teacher in teachers.order_by("name")],
        "periods_json": _builder_periods_data(selected_timetable),
    }
    return render(request, "proxy_adjustment_panel.html", context)


@csrf_exempt
@login_required
@log_exceptions
def proxy_adjustment_data(request):
    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    timetable_id = request.GET.get("timetable_id")
    adjustment_date = _parse_adjustment_date(request.GET.get("date"))
    manual_teacher_id = request.GET.get("teacher_id")

    if not timetable_id:
        return JsonResponse({"success": False, "message": "Timetable is required."})

    timetable = get_object_or_404(Timetable, id=timetable_id, school=current_school)
    statuses = list(TeacherDailyStatus.objects.filter(
        school=timetable.school,
        date=adjustment_date,
    ).select_related("teacher").prefetch_related("unavailable_periods"))
    affected_entries = [
        _entry_adjustment_payload(entry, adjustment_date, statuses, existing_adjustment)
        for entry, existing_adjustment in _affected_proxy_entries(timetable, adjustment_date, statuses, manual_teacher_id)
    ]

    return JsonResponse({
        "success": True,
        "date": adjustment_date.strftime("%Y-%m-%d"),
        "day_name": _date_day_name(adjustment_date),
        "periods": _builder_periods_data(timetable),
        "teachers": [
            _teacher_load_payload(timetable, adjustment_date, teacher)
            for teacher in Teacher.objects.filter(school=timetable.school, timetable=timetable, is_active=True).order_by("name")
        ],
        "teacher_statuses": [_status_data(status) for status in statuses],
        "lectures": affected_entries,
    })


@csrf_exempt
@login_required
@log_exceptions
def auto_proxy_adjustments(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request."})

    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    data = json.loads(request.body)
    timetable = get_object_or_404(Timetable, id=data.get("timetable_id"), school=current_school)
    adjustment_date = _parse_adjustment_date(data.get("date"))
    manual_teacher_id = data.get("teacher_id")
    statuses = list(TeacherDailyStatus.objects.filter(
        school=timetable.school,
        date=adjustment_date,
    ).select_related("teacher").prefetch_related("unavailable_periods"))

    assigned_count = 0
    skipped_count = 0
    already_done_count = 0

    with transaction.atomic():
        for entry, existing_adjustment in _affected_proxy_entries(timetable, adjustment_date, statuses, manual_teacher_id):
            if existing_adjustment and existing_adjustment.status in {"ASSIGNED", "CANCELLED"}:
                already_done_count += 1
                continue

            suggestions = _suggest_proxy_teachers(entry, adjustment_date, statuses)
            if not suggestions:
                skipped_count += 1
                continue

            teacher_status = _teacher_unavailable_status(statuses, entry.teacher_id, entry.period_id_value)
            _create_or_update_adjustment(
                entry,
                adjustment_date,
                {
                    "proxy_teacher_id": suggestions[0]["id"],
                    "status": "ASSIGNED",
                    "reason": data.get("reason", ""),
                    "admin_note": "AI Magic Adjustment",
                },
                teacher_status,
            )
            assigned_count += 1

    return JsonResponse({
        "success": True,
        "message": (
            f"AI Magic Adjustment assigned {assigned_count} lecture(s). "
            f"Skipped {skipped_count} without available proxy. "
            f"{already_done_count} already assigned/cancelled."
        ),
        "assigned_count": assigned_count,
        "skipped_count": skipped_count,
        "already_done_count": already_done_count,
    })


@csrf_exempt
@login_required
@log_exceptions
def save_teacher_daily_status(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request."})

    data = json.loads(request.body)
    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    timetable_id = data.get("timetable_id")
    timetable = get_object_or_404(Timetable, id=timetable_id, school=current_school) if timetable_id else None
    teacher_filter = {
        "id": data.get("teacher_id"),
        "school": current_school,
        "is_active": True,
    }
    if timetable:
        teacher_filter["timetable"] = timetable
    teacher = get_object_or_404(Teacher, **teacher_filter)
    adjustment_date = _parse_adjustment_date(data.get("date"))
    status_type = data.get("status_type") or "LEAVE"
    full_day = bool(data.get("full_day", True))
    period_ids = data.get("period_ids", [])

    if status_type not in {"LEAVE", "IN_SCHOOL_UNAVAILABLE"}:
        return JsonResponse({"success": False, "message": "Invalid status type."})

    status, _ = TeacherDailyStatus.objects.update_or_create(
        teacher=teacher,
        date=adjustment_date,
        defaults={
            "school": teacher.school,
            "status_type": status_type,
            "full_day": full_day,
            "reason": data.get("reason", ""),
            "notes": data.get("notes", ""),
        },
    )

    if full_day:
        status.unavailable_periods.clear()
    else:
        status.unavailable_periods.set(Period.objects.filter(id__in=period_ids))

    return JsonResponse({
        "success": True,
        "message": "Teacher availability status saved.",
        "teacher_status": _status_data(status),
    })


@csrf_exempt
@login_required
@log_exceptions
def delete_teacher_daily_status(request, status_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request."})

    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    status = get_object_or_404(TeacherDailyStatus, id=status_id, school=current_school)
    LectureAdjustment.objects.filter(
        timetable__school=status.school,
        date=status.date,
    ).filter(
        Q(original_teacher=status.teacher) | Q(original_entry__teacher=status.teacher)
    ).delete()
    status.delete()

    return JsonResponse({"success": True, "message": "Teacher status and linked lecture adjustments removed."})


@log_exceptions
def _create_or_update_adjustment(entry, adjustment_date, payload, teacher_status=None):
    return LectureAdjustment.objects.update_or_create(
        date=adjustment_date,
        original_entry=entry,
        defaults={
            "timetable": entry.timetable,
            "teacher_status": teacher_status,
            "original_teacher": entry.teacher,
            "proxy_teacher_id": payload.get("proxy_teacher_id") or None,
            "class_section": entry.class_section,
            "subject": entry.subject,
            "room": entry.room,
            "day_id_value": entry.day_id_value,
            "day_name": entry.day_name,
            "period_id_value": entry.period_id_value,
            "period_name": entry.period_name,
            "status": payload.get("status", "ASSIGNED"),
            "reason": payload.get("reason", ""),
            "admin_note": payload.get("admin_note", ""),
            "is_locked": bool(payload.get("is_locked", False)),
        },
    )


@csrf_exempt
@login_required
@log_exceptions
def save_lecture_adjustment(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request."})

    data = json.loads(request.body)
    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    entry = get_object_or_404(
        TimetableEntry.objects.select_related("timetable", "teacher", "class_section", "subject", "room"),
        id=data.get("entry_id"),
        timetable__school=current_school,
    )
    adjustment_date = _parse_adjustment_date(data.get("date"))
    status_value = data.get("status", "ASSIGNED")
    proxy_teacher_id = data.get("proxy_teacher_id")

    if status_value not in {"PENDING", "ASSIGNED", "CANCELLED"}:
        return JsonResponse({"success": False, "message": "Invalid adjustment status."})

    if status_value == "ASSIGNED":
        if not proxy_teacher_id:
            return JsonResponse({"success": False, "message": "Please select a proxy teacher."})

        proxy_teacher = get_object_or_404(
            Teacher,
            id=proxy_teacher_id,
            school=entry.timetable.school,
            timetable=entry.timetable,
            is_active=True,
        )
        statuses = list(TeacherDailyStatus.objects.filter(
            school=entry.timetable.school,
            date=adjustment_date,
        ).prefetch_related("unavailable_periods"))

        if _teacher_unavailable_status(statuses, proxy_teacher.id, entry.period_id_value):
            return JsonResponse({"success": False, "message": "Selected proxy teacher is unavailable in this period."})

        if not _teacher_static_available(proxy_teacher.id, entry.day_id_value, entry.period_id_value):
            return JsonResponse({"success": False, "message": "Selected proxy teacher is marked unavailable in weekly availability."})

        if _teacher_has_period_conflict(
            entry.timetable,
            adjustment_date,
            proxy_teacher.id,
            entry.day_id_value,
            entry.period_id_value,
            entry.id,
        ):
            return JsonResponse({"success": False, "message": "Selected proxy teacher already has a lecture in this period."})

    teacher_status = TeacherDailyStatus.objects.filter(
        teacher=entry.teacher,
        date=adjustment_date,
    ).first()

    adjustment, _ = _create_or_update_adjustment(entry, adjustment_date, data, teacher_status)

    return JsonResponse({
        "success": True,
        "message": "Lecture adjustment saved.",
        "adjustment": _adjustment_data(adjustment),
    })


@csrf_exempt
@login_required
@log_exceptions
def delete_lecture_adjustment(request, adjustment_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request."})

    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    adjustment = get_object_or_404(LectureAdjustment, id=adjustment_id, timetable__school=current_school)
    adjustment.delete()

    return JsonResponse({"success": True, "message": "Lecture adjustment deleted."})


@login_required
@log_exceptions
def export_proxy_adjustments(request):
    current_school = get_current_school(request)
    if not current_school:
        return JsonResponse({"success": False, "message": "No active school is linked with your session."}, status=403)

    timetable_id = request.GET.get("timetable_id")
    adjustment_date = _parse_adjustment_date(request.GET.get("date"))

    if not timetable_id:
        return JsonResponse({"success": False, "message": "Timetable is required."}, status=400)

    timetable = get_object_or_404(Timetable.objects.select_related("school", "academic_year"), id=timetable_id, school=current_school)
    adjustments = LectureAdjustment.objects.filter(
        timetable=timetable,
        date=adjustment_date,
    ).select_related(
        "class_section",
        "class_section__class_level",
        "class_section__division",
        "subject",
        "original_teacher",
        "proxy_teacher",
    ).order_by("period_id_value", "class_section__class_level__sort_order", "class_section__division__sort_order")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Adjusted Lectures"

    sheet.merge_cells("A1:F1")
    sheet.merge_cells("A2:F2")
    sheet["A1"] = timetable.school.name
    sheet["A2"] = f"Adjusted Lectures - {adjustment_date.strftime('%d-%m-%Y')} ({_date_day_name(adjustment_date)})"

    headers = [
        "Period",
        "Original Teacher",
        "Class-Division",
        "Subject",
        "Adjusted Teacher",
        "Signature",
    ]
    sheet.append([])
    sheet.append(headers)

    for adjustment in adjustments:
        class_level = adjustment.class_section.class_level.name if adjustment.class_section else ""
        division = adjustment.class_section.division.name if adjustment.class_section else ""
        class_division = f"{class_level} {division}".strip()
        teacher_name = adjustment.proxy_teacher.name if adjustment.proxy_teacher else ""

        sheet.append([
            adjustment.period_name,
            adjustment.original_teacher.name if adjustment.original_teacher else "",
            class_division,
            adjustment.subject.name if adjustment.subject else "",
            teacher_name,
            "",
        ])

    school_fill = PatternFill("solid", fgColor="0F172A")
    subtitle_fill = PatternFill("solid", fgColor="DBEAFE")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet["A1"].fill = school_fill
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A2"].fill = subtitle_fill
    sheet["A2"].font = Font(color="0F172A", bold=True, size=12)
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    for row_index in (1, 2):
        for col_index in range(1, 7):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.border = border
            if row_index == 1:
                cell.fill = school_fill
            else:
                cell.fill = subtitle_fill

    for cell in sheet[4]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row_index in range(5, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 32

    widths = [18, 28, 20, 24, 28, 26]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 24
    sheet.freeze_panes = "A5"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"adjusted-lectures-{adjustment_date.strftime('%Y-%m-%d')}.xlsx"
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@log_exceptions
def _safe_sheet_title(title, existing_titles):
    clean = re.sub(r"[\[\]\:\*\?\/\\]", " ", str(title)).strip() or "Timetable"
    clean = re.sub(r"\s+", " ", clean)[:31]
    candidate = clean
    index = 2

    while candidate in existing_titles:
        suffix = f" {index}"
        candidate = f"{clean[:31 - len(suffix)]}{suffix}"
        index += 1

    existing_titles.add(candidate)
    return candidate


@log_exceptions
def _ordered_export_days(periods_data):
    days = []
    seen = set()

    for period in periods_data:
        key = str(period["day_id"])
        if key in seen:
            continue

        seen.add(key)
        days.append({
            "id": period["day_id"],
            "name": period["day_name"],
            "type": period["day_type"],
        })

    return days


@log_exceptions
def _ordered_export_period_rows(periods_data):
    rows = []
    seen = set()

    for period in periods_data:
        key = (
            period["period_number"],
            period["period_name"],
            period["period_type"],
            period["is_teaching_period"],
        )

        if key in seen:
            continue

        seen.add(key)
        rows.append({
            "number": period["period_number"],
            "name": period["period_name"],
            "type": period["period_type"],
            "is_teaching_period": period["is_teaching_period"],
            "start_time": period["start_time"],
            "end_time": period["end_time"],
        })

    return sorted(rows, key=lambda item: item["number"])


@log_exceptions
def _period_for_day_row(periods_data, day_id, row):
    for period in periods_data:
        if (
            str(period["day_id"]) == str(day_id) and
            period["period_number"] == row["number"] and
            period["period_name"] == row["name"] and
            period["period_type"] == row["type"]
        ):
            return period

    return None


@log_exceptions
def _export_entities(scope, timetable):
    configuration = TimetableConfiguration.objects.filter(timetable=timetable).first()
    if scope == "class":
        entities = ClassSection.objects.select_related(
            "class_level",
            "division",
            "class_teacher",
            "default_room",
        ).filter(
            school=timetable.school,
            timetable=timetable,
            is_active=True,
        )
        if configuration:
            configured_section_ids = set(configuration.class_sections.values_list("id", flat=True))
            if configured_section_ids:
                entities = entities.filter(id__in=configured_section_ids)
        return entities.order_by("class_level__sort_order", "division__sort_order", "id")

    entities = Teacher.objects.filter(
        school=timetable.school,
        timetable=timetable,
        is_active=True,
    )
    if configuration:
        configured_teacher_ids = set(configuration.teachers.values_list("id", flat=True))
        if configured_teacher_ids:
            entities = entities.filter(id__in=configured_teacher_ids)
    return entities.order_by("name", "id")


@log_exceptions
def _export_entry_lookup(timetable, scope):
    entries = TimetableEntry.objects.filter(
        timetable=timetable
    ).select_related(
        "class_section",
        "teacher",
        "subject",
        "room",
    )
    lookup = {}

    for entry in entries:
        if scope == "class":
            key = (str(entry.class_section_id), str(entry.day_id_value), str(entry.period_id_value))
        else:
            if not entry.teacher_id:
                continue
            key = (str(entry.teacher_id), str(entry.day_id_value), str(entry.period_id_value))

        lookup[key] = entry

    return lookup


@log_exceptions
def _entry_text(entry, scope):
    if not entry:
        return ""

    parts = []

    if scope == "class":
        parts.append(entry.subject.name if entry.subject else "Subject")
        parts.append(entry.teacher.name if entry.teacher else "Teacher")
    else:
        parts.append(str(entry.class_section) if entry.class_section else "Class")
        parts.append(entry.subject.name if entry.subject else "Subject")

    if entry.room:
        parts.append(f"Room: {entry.room.name}")

    if entry.is_locked:
        parts.append("Locked")

    return "\n".join(parts)


@log_exceptions
def _entity_title(entity, scope):
    if scope == "class":
        return str(entity)

    return entity.name


@log_exceptions
def _entity_subtitle(entity, scope):
    if scope == "class":
        teacher = entity.class_teacher.name if entity.class_teacher else "Not assigned"
        room = entity.default_room.name if entity.default_room else "Not assigned"
        return f"Class Teacher: {teacher} | Default Room: {room}"

    bits = []
    if entity.employee_id:
        bits.append(f"Employee ID: {entity.employee_id}")
    if entity.department:
        bits.append(f"Department: {entity.department}")
    bits.append(f"Max Load: {entity.max_periods_per_day}/day, {entity.max_periods_per_week}/week")
    return " | ".join(bits)


@log_exceptions
def _export_filename(timetable, scope, extension):
    school = re.sub(r"[^A-Za-z0-9]+", "-", timetable.school.short_name or timetable.school.name).strip("-")
    name = re.sub(r"[^A-Za-z0-9]+", "-", timetable.name).strip("-")
    return f"{school}-{name}-{scope}-wise-timetable.{extension}"


@log_exceptions
def _build_excel_timetable(timetable, scope):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    periods_data = _builder_periods_data(timetable)
    days = _ordered_export_days(periods_data)
    rows = _ordered_export_period_rows(periods_data)
    entities = _export_entities(scope, timetable)
    entries = _export_entry_lookup(timetable, scope)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    navy = "0F172A"
    blue = "1D4ED8"
    light_blue = "DBEAFE"
    soft = "F8FAFC"
    border_color = "CBD5E1"
    break_fill = "FFF7ED"
    teaching_fill = "FFFFFF"
    title_fill = PatternFill("solid", fgColor=navy)
    header_fill = PatternFill("solid", fgColor=blue)
    sub_fill = PatternFill("solid", fgColor=light_blue)
    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    @log_exceptions
    def style_heading(sheet, title, subtitle):
        last_col = max(2, len(days) + 1)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)

        sheet["A1"] = timetable.school.name
        sheet["A2"] = title
        sheet["A3"] = subtitle

        for row in range(1, 4):
            cell = sheet.cell(row=row, column=1)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(color="FFFFFF" if row == 1 else navy, bold=True, size=18 if row == 1 else 12)
            cell.fill = title_fill if row == 1 else sub_fill

        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 24
        sheet.row_dimensions[3].height = 34

    summary["A1"] = timetable.school.name
    summary["A2"] = "Timetable Export Summary"
    summary["A4"] = "Timetable"
    summary["B4"] = timetable.name
    summary["A5"] = "Academic Year"
    summary["B5"] = timetable.academic_year.name
    summary["A6"] = "School Code"
    summary["B6"] = timetable.school.school_code
    summary["A7"] = "Contact"
    summary["B7"] = timetable.school.contact_number
    summary["A8"] = "Email"
    summary["B8"] = timetable.school.email
    summary["A9"] = "Export Type"
    summary["B9"] = "Class-wise Timetable" if scope == "class" else "Teacher-wise Timetable"
    summary["A10"] = "Generated On"
    summary["B10"] = timezone.localtime().strftime("%d %b %Y, %I:%M %p")
    summary["A12"] = "Total Sheets"
    summary["B12"] = len(entities)

    summary.merge_cells("A1:D1")
    summary.merge_cells("A2:D2")
    summary["A1"].fill = title_fill
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=20)
    summary["A1"].alignment = Alignment(horizontal="center")
    summary["A2"].fill = sub_fill
    summary["A2"].font = Font(color=navy, bold=True, size=14)
    summary["A2"].alignment = Alignment(horizontal="center")

    for row in range(4, 13):
        summary.cell(row=row, column=1).font = Font(bold=True, color=navy)
        summary.cell(row=row, column=1).fill = PatternFill("solid", fgColor=soft)
        summary.cell(row=row, column=1).border = border
        summary.cell(row=row, column=2).border = border

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 42
    summary.sheet_view.showGridLines = False

    existing_titles = {"Summary"}

    for entity in entities:
        sheet = workbook.create_sheet(_safe_sheet_title(_entity_title(entity, scope), existing_titles))
        sheet.sheet_view.showGridLines = False
        style_heading(sheet, f"{'Class' if scope == 'class' else 'Teacher'} Timetable: {_entity_title(entity, scope)}", _entity_subtitle(entity, scope))

        sheet.cell(row=5, column=1, value="Period / Day")
        for col, day in enumerate(days, start=2):
            sheet.cell(row=5, column=col, value=day["name"])

        for col in range(1, len(days) + 2):
            cell = sheet.cell(row=5, column=col)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_index, period_row in enumerate(rows, start=6):
            label = f"{period_row['name']}\n{period_row['start_time']} - {period_row['end_time']}"
            period_cell = sheet.cell(row=row_index, column=1, value=label)
            period_cell.font = Font(bold=True, color=navy)
            period_cell.fill = sub_fill
            period_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            period_cell.border = border

            for col, day in enumerate(days, start=2):
                period = _period_for_day_row(periods_data, day["id"], period_row)
                cell = sheet.cell(row=row_index, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if not period:
                    cell.fill = PatternFill("solid", fgColor="E5E7EB")
                    continue

                if not period["is_teaching_period"]:
                    cell.value = period["period_name"]
                    cell.fill = PatternFill("solid", fgColor=break_fill)
                    cell.font = Font(bold=True, color="9A3412")
                    continue

                entry = entries.get((str(entity.id), str(period["day_id"]), str(period["period_id"])))
                cell.value = _entry_text(entry, scope)
                cell.fill = PatternFill("solid", fgColor=teaching_fill)
                cell.font = Font(color=navy, bold=bool(entry))

        sheet.freeze_panes = "B6"
        sheet.column_dimensions["A"].width = 22

        for col in range(2, len(days) + 2):
            sheet.column_dimensions[get_column_letter(col)].width = 26

        for row_index in range(6, len(rows) + 6):
            sheet.row_dimensions[row_index].height = 58

        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.35
        sheet.page_margins.bottom = 0.35

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


@log_exceptions
def _build_pdf_timetable(timetable, scope):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    periods_data = _builder_periods_data(timetable)
    days = _ordered_export_days(periods_data)
    rows = _ordered_export_period_rows(periods_data)
    entities = list(_export_entities(scope, timetable))
    entries = _export_entry_lookup(timetable, scope)
    stream = BytesIO()

    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        leftMargin=0.28 * inch,
        rightMargin=0.28 * inch,
        topMargin=0.25 * inch,
        bottomMargin=0.25 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExportTitle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=3,
    )
    subtitle_style = ParagraphStyle(
        "ExportSubtitle",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=9,
        textColor=colors.HexColor("#334155"),
        spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=7,
        leading=8,
        textColor=colors.HexColor("#0F172A"),
    )
    header_style = ParagraphStyle(
        "HeaderCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )
    period_style = ParagraphStyle(
        "PeriodCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
    )

    story = []

    for index, entity in enumerate(entities):
        if index:
            story.append(PageBreak())

        contact_bits = [
            timetable.academic_year.name,
            timetable.name,
            timetable.school.address,
            timetable.school.contact_number,
            timetable.school.email,
        ]
        contact_text = " | ".join([bit for bit in contact_bits if bit])

        story.append(Paragraph(timetable.school.name, title_style))
        story.append(Paragraph(contact_text, subtitle_style))
        story.append(Paragraph(f"{'Class' if scope == 'class' else 'Teacher'} Timetable: {_entity_title(entity, scope)}", title_style))
        story.append(Paragraph(_entity_subtitle(entity, scope), subtitle_style))
        story.append(Spacer(1, 5))

        table_data = [[Paragraph("Period / Day", header_style)] + [Paragraph(day["name"], header_style) for day in days]]

        for period_row in rows:
            label = f"{period_row['name']}<br/>{period_row['start_time']} - {period_row['end_time']}"
            row = [Paragraph(label, period_style)]

            for day in days:
                period = _period_for_day_row(periods_data, day["id"], period_row)

                if not period:
                    row.append("")
                    continue

                if not period["is_teaching_period"]:
                    row.append(Paragraph(f"<b>{period['period_name']}</b>", period_style))
                    continue

                entry = entries.get((str(entity.id), str(period["day_id"]), str(period["period_id"])))
                row.append(Paragraph(_entry_text(entry, scope).replace("\n", "<br/>"), cell_style))

            table_data.append(row)

        usable_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
        first_width = 1.35 * inch
        day_width = (usable_width - first_width) / max(1, len(days))
        table = Table(table_data, colWidths=[first_width] + [day_width] * len(days), repeatRows=1)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#DBEAFE")),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (1, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ])

        for row_index, period_row in enumerate(rows, start=1):
            if not period_row["is_teaching_period"]:
                style.add("BACKGROUND", (1, row_index), (-1, row_index), colors.HexColor("#FFF7ED"))
                style.add("TEXTCOLOR", (1, row_index), (-1, row_index), colors.HexColor("#9A3412"))

        table.setStyle(style)
        story.append(table)

    doc.build(story)
    stream.seek(0)
    return stream.getvalue()


@login_required
@log_exceptions
def export_timetable(request, timetable_id, scope, file_format):
    if scope not in {"class", "teacher"}:
        return JsonResponse({"success": False, "message": "Invalid export scope"}, status=400)

    if file_format not in {"xlsx", "pdf"}:
        return JsonResponse({"success": False, "message": "Invalid export format"}, status=400)

    timetable = get_object_or_404(
        Timetable.objects.select_related("school", "academic_year"),
        id=timetable_id,
    )

    if file_format == "xlsx":
        content = _build_excel_timetable(timetable, scope)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = _build_pdf_timetable(timetable, scope)
        content_type = "application/pdf"

    response = HttpResponse(content, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{_export_filename(timetable, scope, file_format)}"'
    return response
