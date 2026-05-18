import json
import re
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import Teacher
from .forms import TeacherForm
from Schools.models import School
from Subjects.models import TeacherSubjectCapability
from Timetables.models import ClassSection, LessonAllocation, TimetableEntry
from Accounts.utils import get_current_school, get_school_object_or_404, redirect_if_no_current_school, school_queryset
from AI_TIMETABLE_SAAS.logging_utils import log_exceptions


TEACHER_IMPORT_HEADERS = [
    "School Code",
    "School Name",
    "Name",
    "Short Name",
    "Employee ID",
    "Mobile Number",
    "Email",
    "Teacher Type",
    "Max Periods Per Day",
    "Max Periods Per Week",
    "Active",
]


@log_exceptions
def _filtered_teachers(request):
    teachers = school_queryset(
        request,
        Teacher.objects.select_related("school"),
    ).order_by("-id")

    search_query = request.GET.get("search", "")
    teacher_type_filter = request.GET.get("teacher_type", "")
    status_filter = request.GET.get("status", "")

    if search_query:
        teachers = teachers.filter(
            Q(name__icontains=search_query) |
            Q(short_name__icontains=search_query) |
            Q(employee_id__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(school__name__icontains=search_query)
        )

    if teacher_type_filter:
        teachers = teachers.filter(teacher_type=teacher_type_filter)

    if status_filter == "active":
        teachers = teachers.filter(is_active=True)

    if status_filter == "inactive":
        teachers = teachers.filter(is_active=False)

    return teachers, search_query, teacher_type_filter, status_filter


@login_required
@log_exceptions
def teacher_list(request):
    teachers, search_query, teacher_type_filter, status_filter = _filtered_teachers(request)
    current_school = get_current_school(request)

    for teacher in teachers:
        teacher.impact_class_sections = ClassSection.objects.filter(class_teacher=teacher).count()
        teacher.impact_capabilities = TeacherSubjectCapability.objects.filter(teacher=teacher).count()
        teacher.impact_allocations = LessonAllocation.objects.filter(teacher=teacher).count()
        teacher.impact_entries = TimetableEntry.objects.filter(teacher=teacher).count()

    context = {
        "teachers": teachers,
        "total_teachers": teachers.count(),
        "active_teachers": teachers.filter(is_active=True).count(),
        "full_time_teachers": teachers.filter(teacher_type="FULL_TIME").count(),
        "part_time_teachers": teachers.filter(teacher_type="PART_TIME").count(),
        "total_schools": 1 if current_school else 0,
        "search_query": search_query,
        "teacher_type_filter": teacher_type_filter,
        "status_filter": status_filter,
    }

    return render(request, "teacher_list.html", context)


@login_required
@log_exceptions
def teacher_create(request):
    no_school_response = redirect_if_no_current_school(request)
    if no_school_response:
        return no_school_response

    current_school = get_current_school(request)
    if request.method == "POST":
        form = TeacherForm(request.POST, current_school=current_school)

        if form.is_valid():
            form.save()
            messages.success(request, "Teacher created successfully.")
            return HttpResponse("""
            <script>
            window.close();
            </script>
            """)
    else:
        form = TeacherForm(current_school=current_school)
        form.fields["employee_id"].initial = _next_employee_id(current_school)

    return render(request, "teacher_form.html", {
        "form": form,
        "title": "Create Teacher",
        "subtitle": "Add teacher details, contact information and workload limits.",
        "button_text": "Save Teacher",
        "employee_id_prefix": _employee_id_prefix(current_school),
    })


@log_exceptions
def _default_teacher_rows():
    return [
        {
            "name": "",
            "short_name": "",
            "employee_id": "",
            "mobile_number": "",
            "email": "",
            "teacher_type": "FULL_TIME",
            "max_periods_per_day": 6,
            "max_periods_per_week": 30,
            "is_active": True,
        },
        {
            "name": "",
            "short_name": "",
            "employee_id": "",
            "mobile_number": "",
            "email": "",
            "teacher_type": "FULL_TIME",
            "max_periods_per_day": 6,
            "max_periods_per_week": 30,
            "is_active": True,
        },
        {
            "name": "",
            "short_name": "",
            "employee_id": "",
            "mobile_number": "",
            "email": "",
            "teacher_type": "FULL_TIME",
            "max_periods_per_day": 6,
            "max_periods_per_week": 30,
            "is_active": True,
        },
    ]


@log_exceptions
def _employee_id_prefix(school):
    source = school.short_name or school.name or "EMP"
    letters = re.sub(r"[^A-Za-z]", "", source).upper()
    return (letters[:3] or "EMP").ljust(3, "X")


@log_exceptions
def _next_employee_id(school, offset=0):
    prefix = _employee_id_prefix(school)
    existing_ids = Teacher.objects.filter(
        school=school,
        employee_id__istartswith=prefix,
    ).values_list("employee_id", flat=True)

    max_number = 0
    pattern = re.compile(rf"^{re.escape(prefix)}(\d+)$", re.IGNORECASE)
    for employee_id in existing_ids:
        match = pattern.match(str(employee_id or "").strip())
        if match:
            max_number = max(max_number, int(match.group(1)))

    return f"{prefix}{max_number + offset + 1:03d}"


@login_required
@log_exceptions
def teacher_quick_create(request):
    no_school_response = redirect_if_no_current_school(request)
    if no_school_response:
        return no_school_response

    current_school = get_current_school(request)

    if request.method == "POST":
        teacher_rows = json.loads(request.POST.get("teachers_json") or "[]")
        created_count = 0
        updated_count = 0
        skipped_count = 0

        with transaction.atomic():
            for teacher_data in teacher_rows:
                name = str(teacher_data.get("name") or "").strip()
                if not name:
                    skipped_count += 1
                    continue

                teacher_type = str(teacher_data.get("teacher_type") or "FULL_TIME").strip().upper()
                if teacher_type not in {"FULL_TIME", "PART_TIME"}:
                    teacher_type = "FULL_TIME"

                employee_id = str(teacher_data.get("employee_id") or "").strip()
                defaults = {
                    "name": name,
                    "short_name": str(teacher_data.get("short_name") or "").strip(),
                    "employee_id": employee_id,
                    "mobile_number": str(teacher_data.get("mobile_number") or "").strip(),
                    "email": str(teacher_data.get("email") or "").strip(),
                    "teacher_type": teacher_type,
                    "max_periods_per_day": _int_from_excel(teacher_data.get("max_periods_per_day"), 6),
                    "max_periods_per_week": _int_from_excel(teacher_data.get("max_periods_per_week"), 30),
                    "is_active": bool(teacher_data.get("is_active", True)),
                }

                teacher = None
                if employee_id:
                    teacher = Teacher.objects.filter(
                        school=current_school,
                        employee_id__iexact=employee_id,
                    ).first()

                if teacher:
                    for field, value in defaults.items():
                        setattr(teacher, field, value)
                    teacher.save()
                    updated_count += 1
                else:
                    Teacher.objects.create(school=current_school, **defaults)
                    created_count += 1

        messages.success(
            request,
            f"Quick teacher setup completed. Created: {created_count}, Updated: {updated_count}, Skipped blank rows: {skipped_count}."
        )
        return HttpResponse("""
        <script>
        window.close();
        </script>
        """)

    return render(request, "teacher_form.html", {
        "form": TeacherForm(current_school=current_school),
        "title": "AI Auto Generate Teachers",
        "subtitle": "Generate multiple teacher rows quickly, then edit any details before saving.",
        "button_text": "Save Generated Teachers",
        "bulk_create": True,
        "initial_teachers_json": [
            {
                **teacher_row,
                "employee_id": _next_employee_id(current_school, index),
            }
            for index, teacher_row in enumerate(_default_teacher_rows())
        ],
        "employee_id_prefix": _employee_id_prefix(current_school),
    })


@login_required
@log_exceptions
def teacher_update(request, pk):
    teacher = get_school_object_or_404(request, Teacher.objects.all(), pk=pk)
    current_school = get_current_school(request)

    if request.method == "POST":
        form = TeacherForm(request.POST, instance=teacher, current_school=current_school)

        if form.is_valid():
            form.save()
            messages.success(request, "Teacher updated successfully.")
            return HttpResponse("""
            <script>
            window.close();
            </script>
            """)
    else:
        form = TeacherForm(instance=teacher, current_school=current_school)

    return render(request, "teacher_form.html", {
        "form": form,
        "title": "Update Teacher",
        "subtitle": "Update teacher details, workload limits and status.",
        "button_text": "Update Teacher",
    })


@login_required
@require_POST
@log_exceptions
def teacher_delete(request, pk):
    teacher = get_school_object_or_404(request, Teacher.objects.all(), pk=pk)
    name = teacher.name
    teacher.delete()
    messages.success(request, f"Teacher '{name}' deleted successfully.")
    return redirect("teacher_list")


@login_required
@require_POST
@log_exceptions
def teacher_bulk_delete(request):
    selected_ids = request.POST.getlist("teacher_ids")
    if not selected_ids:
        messages.warning(request, "Select at least one teacher to delete.")
        return redirect("teacher_list")

    teachers = school_queryset(
        request,
        Teacher.objects.filter(id__in=selected_ids),
    )
    deleted_count = teachers.count()
    teachers.delete()

    if deleted_count:
        messages.success(request, f"Deleted {deleted_count} selected teacher(s).")
    else:
        messages.info(request, "No teachers were deleted.")

    return redirect("teacher_list")


@log_exceptions
def _teacher_workbook_response(workbook, filename):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@log_exceptions
def _style_teacher_sheet(sheet, title):
    navy = "0F172A"
    blue = "2563EB"
    light_blue = "DBEAFE"
    border_color = "CBD5E1"
    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TEACHER_IMPORT_HEADERS))
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    for col, header in enumerate(TEACHER_IMPORT_HEADERS, start=1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        sheet.column_dimensions[get_column_letter(col)].width = max(16, min(28, len(header) + 6))

    for row in range(4, 504):
        for col in range(1, len(TEACHER_IMPORT_HEADERS) + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False

    teacher_type_validation = DataValidation(type="list", formula1='"FULL_TIME,PART_TIME"', allow_blank=False)
    active_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    sheet.add_data_validation(teacher_type_validation)
    sheet.add_data_validation(active_validation)
    teacher_type_validation.add("H4:H503")
    active_validation.add("K4:K503")

    note = sheet["A2"]
    note.value = "Required: School Code or School Name, Name. Teacher Type must be FULL_TIME or PART_TIME. Active must be TRUE or FALSE."
    note.font = Font(color=navy, italic=True)
    note.alignment = Alignment(wrap_text=True)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(TEACHER_IMPORT_HEADERS))


@login_required
@log_exceptions
def teacher_import_template(request):
    current_school = get_current_school(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Teacher Import"
    _style_teacher_sheet(sheet, "Teacher Bulk Import Template")

    sample_school = current_school or School.objects.filter(is_active=True).order_by("name").first()
    sample_school_code = sample_school.school_code if sample_school else "SCH001"
    sample_school_name = sample_school.name if sample_school else "Sample School"
    sample_rows = [
        [sample_school_code, sample_school_name, "Amit Sharma", "Amit", "EMP001", "9876543210", "amit@example.com", "FULL_TIME", 6, 30, "TRUE"],
        [sample_school_code, sample_school_name, "Priya Patil", "Priya", "EMP002", "9876543211", "priya@example.com", "PART_TIME", 4, 18, "TRUE"],
    ]

    for row_index, row in enumerate(sample_rows, start=4):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    schools_sheet = workbook.create_sheet("Schools")
    schools_sheet.append(["School Code", "School Name"])
    schools = School.objects.filter(pk=current_school.pk) if current_school else School.objects.none()
    for school in schools.order_by("name"):
        schools_sheet.append([school.school_code, school.name])
    schools_sheet.column_dimensions["A"].width = 22
    schools_sheet.column_dimensions["B"].width = 38

    return _teacher_workbook_response(workbook, "teacher-import-template.xlsx")


@login_required
@log_exceptions
def teacher_export_excel(request):
    teachers, _, _, _ = _filtered_teachers(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Teachers"
    _style_teacher_sheet(sheet, "Teacher Export")

    for row_index, teacher in enumerate(teachers, start=4):
        values = [
            teacher.school.school_code,
            teacher.school.name,
            teacher.name,
            teacher.short_name,
            teacher.employee_id,
            teacher.mobile_number,
            teacher.email,
            teacher.teacher_type,
            teacher.max_periods_per_day,
            teacher.max_periods_per_week,
            "TRUE" if teacher.is_active else "FALSE",
        ]
        for col_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    return _teacher_workbook_response(workbook, "teachers-export.xlsx")


@log_exceptions
def _bool_from_excel(value):
    return str(value).strip().upper() in {"TRUE", "YES", "Y", "1", "ACTIVE"}


@log_exceptions
def _int_from_excel(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


@log_exceptions
def _school_from_import(school_code, school_name):
    school_code = str(school_code or "").strip()
    school_name = str(school_name or "").strip()

    if school_code:
        school = School.objects.filter(school_code__iexact=school_code).first()
        if school:
            return school

    if school_name:
        return School.objects.filter(name__iexact=school_name).first()

    return None


@login_required
@require_POST
@log_exceptions
def teacher_import_excel(request):
    current_school = get_current_school(request)
    if not current_school:
        messages.error(request, "No active school is linked with your session.")
        return redirect("teacher_list")

    upload = request.FILES.get("teacher_file")

    if not upload:
        messages.error(request, "Please choose an Excel file to import.")
        return redirect("teacher_list")

    try:
        workbook = load_workbook(upload, data_only=True)
        sheet = workbook.active
    except Exception:
        messages.error(request, "Invalid Excel file. Please download and use the teacher import template.")
        return redirect("teacher_list")

    headers = [str(sheet.cell(row=3, column=col).value or "").strip() for col in range(1, len(TEACHER_IMPORT_HEADERS) + 1)]

    if headers != TEACHER_IMPORT_HEADERS:
        messages.error(request, "Template headers do not match. Please download the latest teacher import template.")
        return redirect("teacher_list")

    created = 0
    updated = 0
    skipped = []

    for row_index in range(4, sheet.max_row + 1):
        values = [sheet.cell(row=row_index, column=col).value for col in range(1, len(TEACHER_IMPORT_HEADERS) + 1)]

        if not any(values):
            continue

        school = _school_from_import(values[0], values[1])
        name = str(values[2] or "").strip()

        if not school or school.id != current_school.id or not name:
            skipped.append(f"Row {row_index}: only your current school and teacher name are allowed")
            continue

        teacher_type = str(values[7] or "FULL_TIME").strip().upper()
        if teacher_type not in {"FULL_TIME", "PART_TIME"}:
            skipped.append(f"Row {row_index}: invalid teacher type")
            continue

        employee_id = str(values[4] or "").strip()
        defaults = {
            "name": name,
            "short_name": str(values[3] or "").strip(),
            "employee_id": employee_id,
            "mobile_number": str(values[5] or "").strip(),
            "email": str(values[6] or "").strip(),
            "teacher_type": teacher_type,
            "max_periods_per_day": _int_from_excel(values[8], 6),
            "max_periods_per_week": _int_from_excel(values[9], 30),
            "is_active": _bool_from_excel(values[10]),
        }

        teacher = None
        if employee_id:
            teacher = Teacher.objects.filter(school=school, employee_id__iexact=employee_id).first()

        if teacher:
            for field, value in defaults.items():
                setattr(teacher, field, value)
            teacher.school = school
            teacher.save()
            updated += 1
        else:
            Teacher.objects.create(school=school, **defaults)
            created += 1

    message = f"Teacher import completed. Created: {created}, Updated: {updated}."
    if skipped:
        message += " Skipped: " + "; ".join(skipped[:5])
        if len(skipped) > 5:
            message += f"; and {len(skipped) - 5} more."
        messages.warning(request, message)
    else:
        messages.success(request, message)

    return redirect("teacher_list")
